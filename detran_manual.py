import time
import re

import os
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

# ================= CONFIGURAÇÕES =================

URL = "https://sistemas.detran.ce.gov.br/central"
EXCEL_ARQUIVO = "resultado_detran_organizado.xlsx"
INTERVALO_ENTRE_CONSULTAS = 2  # segundos - reduzido de 5

VEICULOS = [
    {"placa": "SBA7F09", "renavam": "01365705622"},
    {"placa": "TIF1J98", "renavam": "01450499292"},
]

TIMEOUT_PADRAO = 20000
TIMEOUT_MULTAS = 20000
TIMEOUT_TABELA = 20000

DELAY_SCROLL = 0.2  # reduzido de 0.4
DELAY_CHECKBOX = 0.2  # reduzido de 0.4
DELAY_EMITIR = 2  # reduzido de 4
DELAY_DIGITACAO = 0.1  # reduzido de 0.3

REGEX_BOTAO_CONSULTAR = re.compile("consultar|confirmar|pesquisar", re.I)
REGEX_BOTAO_FECHAR = re.compile("fechar", re.I)
REGEX_BOTAO_EMITIR = re.compile("emitir", re.I)
REGEX_CLIQUE_AQUI = re.compile("clique aqui", re.I)
REGEX_VALOR = re.compile(r"R\$[\s]*([\d.,]+)")
REGEX_MULTAS = re.compile(r"possui\s+(\d+)\s+multa", re.I)

# ================= UTIL =================

def log(msg):
    print(msg)

def formatar_valor_br(valor):
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# ================= FORM =================

def preencher_dados(page, placa, renavam):
    """Preenche placa e renavam com delay entre caracteres"""
    campo_placa = page.locator('input[placeholder*="Placa" i]')
    campo_renavam = page.locator('input[placeholder*="Renavam" i]')
    
    # Limpa e preenche placa com delay
    campo_placa.click(force=True)
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    for char in placa:
        page.keyboard.press(char)
        time.sleep(DELAY_DIGITACAO)
    
    # Limpa e preenche renavam com delay
    campo_renavam.click(force=True)
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    for char in renavam:
        page.keyboard.press(char)
        time.sleep(DELAY_DIGITACAO)

# ================= AÇÕES =================

def fechar_popup(page):
    try:
        page.get_by_role("button", name=REGEX_BOTAO_FECHAR).click(timeout=3000)
    except:
        pass

def acessar_taxas_multas(page):
    page.get_by_text("Taxas / Multas", exact=False).click()

def clicar_consultar(page):
    with page.expect_navigation(wait_until="networkidle"):
        page.get_by_role("button", name=REGEX_BOTAO_CONSULTAR).click()

# ================= MULTAS =================

def abrir_detalhe_multas(page):
    page.get_by_text(REGEX_CLIQUE_AQUI).first.wait_for(timeout=TIMEOUT_MULTAS)
    page.get_by_text(REGEX_CLIQUE_AQUI).first.click()
    page.wait_for_load_state("networkidle")
    log("🔍 Tela de multas aberta")

def extrair_valor(texto):
    valores = REGEX_VALOR.findall(texto)
    if valores:
        return float(valores[-1].replace(".", "").replace(",", "."))
    return 0.0

def processar_multas(page):
    tabela = page.locator("table")
    tabela.wait_for(timeout=TIMEOUT_TABELA)

    linhas = tabela.locator("tbody tr")
    qtd = linhas.count()

    indices_validos = []
    total = 0.0
    motivos = []

    for i in range(qtd):
        linha = linhas.nth(i)
        texto = linha.inner_text().replace("\n", " ")
        valor = extrair_valor(texto)

        if valor > 0:
            indices_validos.append(i)
            total += valor
            motivos.append(texto)
            log(f"📝 Multa válida linha {i} → R$ {valor:.2f}")

    log(f"💰 Total calculado: R$ {formatar_valor_br(total)}")
    return motivos, total, indices_validos

# ================= SELEÇÃO CORRETA DAS MULTAS =================

def marcar_checkboxes_multas(page, indices):
    tabela = page.locator("table")
    linhas = tabela.locator("tbody tr")

    marcadas = 0

    for i in indices:
        linha = linhas.nth(i)
        linha.scroll_into_view_if_needed()
        time.sleep(DELAY_SCROLL)

        try:
            # 🔥 CLICA NO ELEMENTO REAL DO CHECKBOX (Material UI)
            checkbox = linha.locator(
                'mat-checkbox label, mat-checkbox span, input[type="checkbox"]'
            ).first

            checkbox.click(force=True)
            time.sleep(DELAY_CHECKBOX)
            marcadas += 1
            log(f"☑️ Multa {marcadas} selecionada (linha {i})")

        except Exception as e:
            log(f"⚠️ Falha ao marcar linha {i}: {e}")

    log(f"✅ {marcadas} multas selecionadas com sucesso")

def extrair_codigo_pix(page):
    """Extrai o código de pagamento PIX da página antes de emitir."""
    try:
        # Procura pelo botão com onclick="copiarParaClipboard('pix-multas')"
        # ou similar e extrai o valor associado
        
        # Tenta encontrar o elemento com o atributo onclick
        elementos = page.locator('[onclick*="pix"]').all() if page.locator('[onclick*="pix"]').count() > 0 else []
        
        if elementos:
            for elem in elementos:
                texto = elem.inner_text() if elem else ""
                log(f"🔍 Elemento PIX encontrado: {texto}")
        
        # Tenta extrair código de pagamento do texto visível
        texto_pagina = page.inner_text("body")
        
        # Procura por padrão de código de pagamento: números separados por espaço
        # Formato típico: 856300000010 041300062027 601302026898 06128693005
        padrao_codigo = r"(\d{12}\s+\d{12}\s+\d{12}\s+\d{11})"
        match = re.search(padrao_codigo, texto_pagina)
        
        if match:
            codigo = match.group(1).strip()
            log(f"💳 Código PIX extraído: {codigo}")
            return codigo
        
        log("⚠️ Código PIX não encontrado na página")
        return "-"
    except Exception as e:
        log(f"⚠️ Erro ao extrair código PIX: {e}")
        return "-"

    log(f"✅ {marcadas} multas selecionadas com sucesso")

def clicar_emitir(page, context, pasta_boletos):
    """Clica em Emitir, espera aparecer o botão Baixar Extrato e baixa o PDF."""
    botao_emitir = page.get_by_role("button", name=REGEX_BOTAO_EMITIR)
    botao_emitir.wait_for(timeout=TIMEOUT_TABELA)

    def salvar_download(download):
        nome_arquivo = download.suggested_filename or f"extrato_{int(time.time())}.pdf"
        caminho_destino = os.path.join(pasta_boletos, nome_arquivo)
        download.save_as(caminho_destino)
        log(f"💾 Boleto salvo via download: {caminho_destino}")
        return caminho_destino

    # 1) Clica em Emitir para revelar o botão "Baixar Extrato"
    botao_emitir.click()
    log("🧾 Emitir clicado")
    page.wait_for_timeout(800)

    # Localiza o botão Baixar Extrato (ou variações) mostrado na imagem
    seletor_baixar = (
        'button:has-text("Baixar Extrato"), a:has-text("Baixar Extrato"), '
        'button:has-text("Baixar"), a:has-text("Baixar"), '
        'button:has-text("Extrato"), a:has-text("Extrato")'
    )
    botao_baixar = page.locator(seletor_baixar).first

    try:
        botao_baixar.wait_for(timeout=20000)
    except Exception:
        log("⚠️ Botão Baixar Extrato não apareceu.")
        return None

    # 2) Clica em Baixar Extrato - isso abre o PDF em nova aba
    botao_baixar.click(force=True)
    log("⬇️ Baixar Extrato clicado")
    page.wait_for_timeout(2000)

    # 3) Captura a nova página/aba que abriu com o PDF
    paginas = context.pages
    pagina_pdf = None
    
    for p in reversed(paginas):
        if "gerar_boleto" in p.url or "pdf" in p.url.lower():
            pagina_pdf = p
            break
    
    if not pagina_pdf:
        log("⚠️ Nenhuma aba PDF encontrada")
        return None
    
    log(f"📄 PDF aberto em nova aba")
    pagina_pdf.wait_for_load_state("load", timeout=15000)
    page.wait_for_timeout(3000)
    
    # 4) Clica no ícone de download no viewer do PDF
    try:
        # Procura especificamente pelo botão "Baixar Extrato" dentro da página
        seletores_download = [
            'button#btn-exibir-extrato',  # ID específico do botão
            'button.btn.btn-success#btn-exibir-extrato',  # Combinação de classe e ID
            'button[id="btn-exibir-extrato"]',  # Seletor alternativo
            'button[aria-label="Fazer download"]',
            'button[aria-label="Download"]',
            '#download',
            'button#download',
            'cr-icon-button#download',
            'button[aria-label*="download" i]',
            'button[title*="download" i]',
            'button[title*="Download" i]',
            '[role="button"][aria-label*="download" i]',
        ]
        
        log("🔍 Procurando botão de download...")
        botao_download_encontrado = False
        
        for seletor in seletores_download:
            try:
                botao = pagina_pdf.locator(seletor).first
                if botao.is_visible(timeout=1000):
                    log(f"✅ Encontrou botão com seletor: {seletor}")
                    botao.click(force=True)
                    log("✅ Clicou no botão de download")
                    botao_download_encontrado = True
                    page.wait_for_timeout(1000)
                    break
            except Exception as e:
                pass
        
        if not botao_download_encontrado:
            log("⚠️ Botão visual não encontrado, tentando Ctrl+S...")
            pagina_pdf.keyboard.press("Control+S")
            page.wait_for_timeout(1500)
        
        # Aguarda o download
        try:
            with pagina_pdf.expect_download(timeout=25000) as download_info:
                page.wait_for_timeout(2000)
            
            download = download_info.value
            nome_arquivo = download.suggested_filename or f"extrato_{int(time.time())}.pdf"
            caminho_destino = os.path.join(pasta_boletos, nome_arquivo)
            download.save_as(caminho_destino)
            log(f"💾 PDF salvo: {caminho_destino}")
            
            pagina_pdf.close()
            return caminho_destino
        except TimeoutError:
            log("⚠️ Timeout esperando download")
            pagina_pdf.close()
            return None
        
    except Exception as e:
        log(f"⚠️ Erro ao tentar baixar PDF: {e}")
        try:
            pagina_pdf.close()
        except:
            pass
        return None

def extrair_dados_do_pdf(caminho_pdf):
    """Extrai código de pagamento, órgão autuador, descrição e datas do PDF."""
    try:
        if not pdfplumber:
            log("⚠️ pdfplumber não está instalado")
            return "-", "-", "-", "-"
        
        # Valida se o arquivo existe e é PDF
        if not os.path.exists(caminho_pdf):
            log(f"⚠️ Arquivo não encontrado: {caminho_pdf}")
            return "-", "-", "-", "-"
        
        with open(caminho_pdf, 'rb') as f:
            header = f.read(10)
            if not header.startswith(b'%PDF'):
                log(f"⚠️ Arquivo {caminho_pdf} não é um PDF válido")
                return "-", "-", "-", "-"
        
        with pdfplumber.open(caminho_pdf) as pdf:
            texto = ""
            linhas = []
            for page in pdf.pages[:2]:  # Lê primeiras 2 páginas (cabeçalho e descrição)
                conteudo = page.extract_text() or ""
                texto += conteudo
                linhas.extend(conteudo.splitlines())

            log("🔎 Prévia do PDF (linhas iniciais):")
            for l in linhas[:8]:
                log(f"   {l}")

            codigo_pagamento = "-"
            descricao_pdf = "-"
            orgao = "-"
            data_infracao_pdf = "-"
            vencimento_pdf = "-"

            # 1) Extrai código de pagamento - procura por padrão numérico específico
            # Geralmente tem 47-48 dígitos em grupos separados por espaços
            for i, linha in enumerate(linhas):
                linha_limpa = linha.strip()
                apenas_digitos = re.sub(r"\D", "", linha_limpa)
                
                # Código de barras tem 47-48 dígitos e geralmente está em linha própria
                # Não deve conter texto além de números e espaços
                if len(apenas_digitos) >= 47 and len(apenas_digitos) <= 48:
                    # Verifica se linha tem pouco texto além de números (evita linhas com descrição)
                    if len(linha_limpa.replace(" ", "")) == len(apenas_digitos):
                        codigo_pagamento = linha_limpa
                        log(f"💳 Código de Pagamento encontrado: {codigo_pagamento}")
                        break
            
            # 2) Extrai órgão autuador - NOVA ABORDAGEM: pega da linha da multa
            # Procura pela linha que contém DETRAN/DEMUTRAN | código | descrição
            for i, linha in enumerate(linhas):
                if ("DETRAN" in linha or "DEMUTRAN" in linha) and "|" in linha:
                    # Extrai o órgão que está antes do primeiro "|"
                    match_orgao = re.match(r"^([^|]+)", linha)
                    if match_orgao:
                        orgao = match_orgao.group(1).strip()
                        log(f"🏢 Órgão Autuador encontrado (linha da multa): {orgao}")
                        break
            
            # FALLBACK: Se não encontrou na linha da multa, procura por padrões
            if orgao == "-":
                # Procura por padrões de órgãos específicos
                padrao_orgaos = [
                    (r"DEMUTRAN\s+[A-Z]+", "DEMUTRAN"),
                    (r"DETRAN-[A-Z]{2}", "DETRAN"),
                    (r"SEMOB", "SEMOB"),
                    (r"POL[IÍ]CIA\s+MILITAR", "PM"),
                    (r"POL[IÍ]CIA\s+FEDERAL", "PF"),
                    (r"POL[IÍ]CIA\s+RODOVI[ÁA]RIA", "PRF"),
                ]
                
                for pattern, fallback in padrao_orgaos:
                    match = re.search(pattern, texto, re.IGNORECASE)
                    if match:
                        orgao = match.group(0).strip()
                        log(f"🏢 Órgão Autuador encontrado (padrão): {orgao}")
                        break
            
            # 3) Extrai descrição: pega a linha logo após "Descrição (Taxa / Multa)"
            for i, linha in enumerate(linhas):
                linha_low = linha.lower()
                if "descri" in linha_low and "taxa" in linha_low:
                    for proxima in linhas[i+1:]:
                        proxima_limpa = proxima.strip()
                        if proxima_limpa:
                            descricao_pdf = proxima_limpa
                            break
                    break
            
            # 4) Extrai datas - procura especificamente pela linha da multa com as duas datas
            datas_encontradas = re.findall(r"\d{2}/\d{2}/\d{4}", texto)
            log(f"📅 Datas encontradas no PDF: {datas_encontradas}")
            
            # Exibe contexto das linhas para debug
            log("📄 Linhas do PDF (primeiras 50):")
            for idx, l in enumerate(linhas[:50]):
                log(f"   [{idx}] {l}")
            
            # MÉTODO PRINCIPAL: Procura pela linha com DETRAN, código da infração e as 2 datas
            # Exemplo: DETRAN-CE | V607910965 | 07455 | TRANSITAR EM VELOCIDADE 06/11/2025 30/01/2026 130,16 104,13
            data_infra_encontrada = False
            vencimento_encontrado = False
            
            for i, linha in enumerate(linhas):
                linha_strip = linha.strip()
                
                # Procura por linha que contenha padrão de multa DETRAN-CE | código | descrição + duas datas
                if ("DETRAN" in linha or "DEMUTRAN" in linha or "|" in linha) and re.search(r"\d{2}/\d{2}/\d{4}", linha):
                    # Encontra todas as datas nesta linha específica
                    datas_na_linha = re.findall(r"\d{2}/\d{2}/\d{4}", linha)
                    
                    if len(datas_na_linha) >= 2:
                        # A primeira data é a infração, a segunda é o vencimento
                        data_infracao_pdf = datas_na_linha[0]
                        vencimento_pdf = datas_na_linha[1]
                        
                        log(f"✅ LINHA DA MULTA ENCONTRADA [{i}]: {linha_strip}")
                        log(f"✅ Data Infração: {data_infracao_pdf}")
                        log(f"✅ Vencimento: {vencimento_pdf}")
                        
                        data_infra_encontrada = True
                        vencimento_encontrado = True
                        break
            
            # MÉTODO ALTERNATIVO 1: Se não encontrou na linha da multa, procura pelos cabeçalhos
            if not data_infra_encontrada or not vencimento_encontrado:
                log("⚠️  Método principal não encontrou. Tentando método alternativo com cabeçalhos...")
                
                for i, linha in enumerate(linhas):
                    linha_low = linha.lower().strip()
                    
                    # Procura pelo cabeçalho da tabela: "Descrição ... Data Infração Vencimento"
                    if "data" in linha_low and "infra" in linha_low and "venci" in linha_low:
                        log(f"🔍 Cabeçalho da tabela encontrado na linha {i}: '{linha}'")
                        
                        # A linha seguinte deve conter os dados da multa
                        for j in range(1, 6):
                            if i+j < len(linhas):
                                proxima = linhas[i+j]
                                datas_na_linha = re.findall(r"\d{2}/\d{2}/\d{4}", proxima)
                                
                                # Filtra datas que não são emissão/processamento (geralmente 2025/2026)
                                if len(datas_na_linha) >= 2:
                                    data_infracao_pdf = datas_na_linha[0]
                                    vencimento_pdf = datas_na_linha[1]
                                    
                                    log(f"✅ Dados encontrados na linha +{j}: {proxima.strip()}")
                                    log(f"✅ Data Infração: {data_infracao_pdf}")
                                    log(f"✅ Vencimento: {vencimento_pdf}")
                                    
                                    data_infra_encontrada = True
                                    vencimento_encontrado = True
                                    break
                        
                        if data_infra_encontrada:
                            break
            
            # MÉTODO ALTERNATIVO 2: Usa lógica de ordenação e filtragem de datas
            if not data_infra_encontrada or not vencimento_encontrado:
                log("⚠️  Métodos anteriores falharam. Usando lógica de ordenação...")
                
                if len(datas_encontradas) >= 2:
                    try:
                        # Remove datas de emissão/geração (geralmente a mais recente e a de hoje)
                        # E remove datas muito antigas (leis/normas)
                        datas_validas = []
                        hoje = datetime.now()
                        
                        for d in datas_encontradas:
                            try:
                                dt = datetime.strptime(d, "%d/%m/%Y")
                                # Filtra datas entre 2020 e 2030 (período válido para multas)
                                if 2020 <= dt.year <= 2030:
                                    datas_validas.append((d, dt))
                            except:
                                pass
                        
                        # Ordena por data
                        datas_validas.sort(key=lambda x: x[1])
                        log(f"📅 Datas válidas ordenadas: {[d[0] for d in datas_validas]}")
                        
                        if len(datas_validas) >= 2:
                            # Infração geralmente é a data mais antiga (quando ocorreu)
                            # Vencimento é posterior
                            data_infracao_pdf = datas_validas[0][0]
                            
                            # Vencimento: procura uma data que seja posterior à infração
                            for d, dt in datas_validas[1:]:
                                if dt > datas_validas[0][1]:
                                    vencimento_pdf = d
                                    break
                            
                            log(f"🔄 Data Infração: {data_infracao_pdf}")
                            log(f"🔄 Vencimento: {vencimento_pdf}")
                    
                    except Exception as e:
                        log(f"❌ Erro no método de ordenação: {e}")

            # Fallback final: se ainda não encontrou, usa últimas datas disponíveis
            if (data_infracao_pdf == "-" or vencimento_pdf == "-") and len(datas_encontradas) >= 2:
                if data_infracao_pdf == "-":
                    data_infracao_pdf = datas_encontradas[0]
                if vencimento_pdf == "-":
                    vencimento_pdf = datas_encontradas[-1]
                log(f"📅 Datas determinadas por fallback: Infração={data_infracao_pdf}, Vencimento={vencimento_pdf}")

            # 5) Combina código de pagamento + descrição na variável final
            resultado_pdf = descricao_pdf
            if codigo_pagamento != "-":
                if descricao_pdf != "-":
                    resultado_pdf = f"{codigo_pagamento} | {descricao_pdf}"
                else:
                    resultado_pdf = codigo_pagamento

        return orgao, resultado_pdf, data_infracao_pdf, vencimento_pdf
    except Exception as e:
        log(f"⚠️ Erro ao ler PDF: {e}")
        return "-", "-", "-", "-"

def reprocessar_pdfs_e_atualizar_excel():
    """Reprocessa todos os PDFs existentes e atualiza o Excel"""
    log("\n🔄 REPROCESSANDO PDFs EXISTENTES...")
    
    # Verifica se existe Excel
    if not os.path.exists(EXCEL_ARQUIVO):
        log(f"❌ Arquivo {EXCEL_ARQUIVO} não encontrado!")
        return
    
    # Carrega Excel atual
    try:
        df = pd.read_excel(EXCEL_ARQUIVO, engine='openpyxl')
    except Exception as e:
        log(f"❌ Erro ao ler Excel: {e}")
        return
    
    log(f"📊 Excel carregado: {len(df)} multas")
    
    # Mapeia PDFs por placa
    pdfs_encontrados = {}
    pasta_boletos = "boletos"
    
    if not os.path.exists(pasta_boletos):
        log(f"❌ Pasta {pasta_boletos} não encontrada!")
        return
    
    # Busca todos os PDFs
    for subpasta in os.listdir(pasta_boletos):
        caminho_subpasta = os.path.join(pasta_boletos, subpasta)
        if os.path.isdir(caminho_subpasta):
            for arquivo in os.listdir(caminho_subpasta):
                if arquivo.endswith('.pdf'):
                    caminho_completo = os.path.join(caminho_subpasta, arquivo)
                    # Tenta extrair placa do nome do arquivo (Extrato_6601163057.pdf)
                    # Na verdade, vamos processar todos os PDFs e associar pela data
                    pdfs_encontrados[caminho_completo] = None
    
    log(f"📄 Encontrados {len(pdfs_encontrados)} PDFs")
    
    # Contador de atualizações
    atualizados = 0
    
    # Para cada PDF, extrai dados
    for caminho_pdf in pdfs_encontrados.keys():
        log(f"\n📑 Processando: {os.path.basename(caminho_pdf)}")
        
        orgao, codigo_barras, data_infracao, data_vencimento = extrair_dados_do_pdf(caminho_pdf)
        
        if orgao == "-" and codigo_barras == "-":
            log(f"⚠️ Nenhum dado extraído de {os.path.basename(caminho_pdf)}")
            continue
        
        log(f"   Órgão: {orgao}")
        log(f"   Código: {codigo_barras[:50]}..." if len(codigo_barras) > 50 else f"   Código: {codigo_barras}")
        log(f"   Data Infração: {data_infracao}")
        log(f"   Data Vencimento: {data_vencimento}")
        
        # Procura no Excel pela data de vencimento ou data de infração
        # Como não temos identificador único, vamos atualizar todas as linhas com órgão vazio
        # e que tenham datas próximas ou vazias
        
        for idx in df.index:
            # Se órgão já está preenchido, pula
            if pd.notna(df.loc[idx, "Órgão Autuador"]) and df.loc[idx, "Órgão Autuador"] != "-":
                continue
            
            # Se código já está preenchido, pula
            if pd.notna(df.loc[idx, "Código de pagamento em barra"]) and df.loc[idx, "Código de pagamento em barra"] != "-":
                continue
            
            # Atualiza primeira linha vazia encontrada
            df.loc[idx, "Órgão Autuador"] = orgao
            df.loc[idx, "Código de pagamento em barra"] = codigo_barras
            
            if data_infracao != "-":
                df.loc[idx, "Data Infração"] = data_infracao
            if data_vencimento != "-":
                df.loc[idx, "Data Vencimento"] = data_vencimento
            
            atualizados += 1
            log(f"   ✅ Atualizado linha {idx + 2}")  # +2 porque índice começa em 0 e tem cabeçalho
            break  # Atualiza apenas 1 linha por PDF
    
    # Salva Excel atualizado
    if atualizados > 0:
        try:
            df.to_excel(EXCEL_ARQUIVO, index=False, sheet_name="Resultado DETRAN", engine='openpyxl')
            log(f"\n✅ Excel atualizado com sucesso! {atualizados} multas atualizadas")
            
            # Aplica formatação
            try:
                wb = openpyxl.load_workbook(EXCEL_ARQUIVO)
                ws = wb.active
                
                # Formata cabeçalho
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Bordas
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin_border
                
                # Congela primeira linha
                ws.freeze_panes = "A2"
                
                # Ajusta largura
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column].width = min(max_length + 2, 50)
                
                wb.save(EXCEL_ARQUIVO)
                log("✅ Formatação aplicada")
            except Exception as e:
                log(f"⚠️ Erro ao formatar Excel: {e}")
        except Exception as e:
            log(f"❌ Erro ao salvar Excel: {e}")
    else:
        log("\n⚠️ Nenhuma multa foi atualizada")

# ================= PROCESSAMENTO =================

def extrair_pendencias(texto):
    match = REGEX_MULTAS.search(texto)
    return int(match.group(1)) if match else 0

def salvar_no_excel(multas_lista):
    """Salva multas no Excel com formatação"""
    if not multas_lista:
        log("⚠️ Nenhuma multa para salvar")
        return
    
    # Define a ordem correta das colunas
    colunas_ordem = [
        "Placa", "#", "AIT", "AIT Originária", "Motivo", 
        "Data Infração", "Data Vencimento", "Valor", "Valor a Pagar", 
        "Órgão Autuador", "Código de pagamento em barra"
    ]
    
    df_novo = pd.DataFrame(multas_lista)
    
    # Reordena as colunas para garantir ordem correta
    # Inclui apenas colunas que existem no DataFrame
    colunas_existentes = [col for col in colunas_ordem if col in df_novo.columns]
    df_novo = df_novo[colunas_existentes]
    
    try:
        # Tenta fechar arquivo se estiver aberto
        import os
        if os.path.exists(EXCEL_ARQUIVO):
            try:
                import gc
                gc.collect()
            except:
                pass
        
        # Salva o novo DataFrame
        df_novo.to_excel(EXCEL_ARQUIVO, index=False, sheet_name="Resultado DETRAN", engine='openpyxl')
    except PermissionError:
        log(f"⚠️ Arquivo {EXCEL_ARQUIVO} está aberto. Feche e tente novamente!")
        return
    except Exception as e:
        log(f"⚠️ Erro ao salvar Excel: {e}")
        return
    
    # Formatar Excel
    try:
        wb = openpyxl.load_workbook(EXCEL_ARQUIVO)
        ws = wb.active
        
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                # Coluna E (Motivo) e K (Código barras) ficam alinhadas à esquerda
                cell.alignment = left if cell.column in (5, 11) else center
        
        # Define largura das colunas
        larguras = {
            "A": 12,  # Placa
            "B": 5,   # #
            "C": 15,  # AIT
            "D": 18,  # AIT Originária
            "E": 55,  # Motivo
            "F": 14,  # Data Infração
            "G": 14,  # Data Vencimento
            "H": 16,  # Valor
            "I": 16,  # Valor a Pagar
            "J": 18,  # Órgão Autuador
            "K": 55   # Código de pagamento em barra
        }
        for col, w in larguras.items():
            ws.column_dimensions[col].width = w
        
        ws.freeze_panes = "A2"
        wb.save(EXCEL_ARQUIVO)
        log(f"✅ Dados salvos em: {EXCEL_ARQUIVO}")
    except Exception as e:
        log(f"⚠️ Erro ao formatar Excel: {e}")

def processar_veiculo(browser, veiculo, indice):
    log("\n" + "=" * 50)
    log(f"🚗 CONSULTA {indice} | {veiculo['placa']}")

    # Cria pasta de download com data de hoje
    pasta_base = "boletos"
    data_hoje = datetime.now().strftime("%d-%m-%Y")
    pasta_boletos = os.path.join(pasta_base, data_hoje)
    
    if not os.path.exists(pasta_boletos):
        os.makedirs(pasta_boletos)
        log(f"📁 Pasta '{pasta_boletos}' criada")

    context = browser.new_context(
        accept_downloads=True
    )
    page = context.new_page()
    multas_lista = []
    numero_sequencial = 0

    try:
        page.goto(URL)
        fechar_popup(page)
        acessar_taxas_multas(page)
        preencher_dados(page, veiculo["placa"], veiculo["renavam"])
        clicar_consultar(page)

        texto = page.inner_text("body").lower()
        qtd_multas = extrair_pendencias(texto)

        log(f"📄 Multas encontradas: {qtd_multas}")

        total = 0.0
        motivos = []

        if qtd_multas > 0:
            abrir_detalhe_multas(page)
            motivos, total, indices = processar_multas(page)
            
            # Processa cada multa para salvar no Excel
            for motivo in motivos:
                numero_sequencial += 1
                
                # DEBUG: Mostra o texto bruto
                log(f"\n🔍 TEXTO BRUTO MULTA {numero_sequencial}:")
                log(f"  {motivo[:200]}...")
                
                # Extrai AIT
                ait = "-"
                match_ait = re.search(r"([A-Z]{1,3}\d{6,})\s*--", motivo)
                if match_ait:
                    ait = match_ait.group(1)
                
                # Extrai datas
                datas = re.findall(r"\d{2}/\d{2}/\d{4}", motivo)
                log(f"  🔍 Datas encontradas (ordem): {datas}")
                
                # Geralmente vem: [vencimento, data_infracao] - vamos inverter
                if len(datas) >= 2:
                    # Assumindo que a primeira data é o vencimento e a segunda é a infração
                    # Se a primeira data for MAIOR que a segunda, está correto
                    # Senão, inverte
                    try:
                        data1 = datetime.strptime(datas[0], "%d/%m/%Y")
                        data2 = datetime.strptime(datas[1], "%d/%m/%Y")
                        
                        # Se data1 > data2, então data1 é vencimento e data2 é infração
                        if data1 > data2:
                            vencimento = datas[0]
                            data_infracao = datas[1]
                        else:
                            # Senão, assume ordem normal
                            data_infracao = datas[0]
                            vencimento = datas[1]
                    except:
                        # Se falhar o parse, usa ordem padrão
                        data_infracao = datas[0]
                        vencimento = datas[1]
                elif len(datas) == 1:
                    data_infracao = datas[0]
                    vencimento = "-"
                else:
                    data_infracao = "-"
                    vencimento = "-"
                
                # Extrai valores
                valores = re.findall(r"R\$\s*([\d.,]+)", motivo)
                valor = "-"
                valor_a_pagar = "-"
                if len(valores) == 1:
                    valor = f"R$ {valores[0]}"
                    valor_a_pagar = f"R$ {valores[0]}"
                elif len(valores) >= 2:
                    valor = f"R$ {valores[-2]}"
                    valor_a_pagar = f"R$ {valores[-1]}"
                
                # Extrai descrição - versão SIMPLIFICADA
                # Remove checkbox, AIT, datas e valores
                descricao = motivo
                # Remove checkbox vazio no início
                descricao = re.sub(r"^\s*\□?\s*", "", descricao)
                # Remove AIT
                descricao = re.sub(r"[A-Z]{1,3}\d{6,}\s*--\s*", "", descricao)
                # Remove datas
                descricao = re.sub(r"\d{2}/\d{2}/\d{4}", "", descricao)
                # Remove valores
                descricao = re.sub(r"R\$\s*[\d.,]+", "", descricao)
                # Remove espaços extras
                descricao = re.sub(r"\s+", " ", descricao).strip()
                
                if not descricao:
                    descricao = "-"
                
                # Exibe informações da multa
                log(f"\n✏️ MULTA {numero_sequencial}")
                log(f"  AIT: {ait}")
                log(f"  📋 Descrição: {descricao}")
                log(f"  📅 Data: {data_infracao} | Vencimento: {vencimento}")
                log(f"  💰 Valor: {valor} → A Pagar: {valor_a_pagar}")
                
                multas_lista.append({
                    "Placa": veiculo["placa"],
                    "#": numero_sequencial,
                    "AIT": ait,
                    "AIT Originária": "-",
                    "Motivo": descricao,
                    "Data Infração": data_infracao,
                    "Data Vencimento": vencimento,
                    "Valor": valor,
                    "Valor a Pagar": valor_a_pagar,
                    "Órgão Autuador": "-",
                    "Código de pagamento em barra": "-"
                })
            
            marcar_checkboxes_multas(page, indices)
            
            # Extrai o código PIX ANTES de emitir
            codigo_pix = extrair_codigo_pix(page)
            
            # Emite, baixa o PDF e extrai dados
            orgao_autuador = "-"
            descricao_pdf = "-"
            data_infracao_pdf = "-"
            vencimento_pdf = "-"
            caminho_pdf = clicar_emitir(page, context, pasta_boletos)
            if caminho_pdf:
                orgao_autuador, descricao_pdf, data_infracao_pdf, vencimento_pdf = extrair_dados_do_pdf(caminho_pdf)
                log(f"🏢 Órgão Autuador: {orgao_autuador}")
                log(f"📄 Descrição PDF: {descricao_pdf}")
                log(f"📅 Datas do PDF - Infração: {data_infracao_pdf}, Vencimento: {vencimento_pdf}")

            # Adiciona código PIX na descrição se encontrou
            if codigo_pix != "-":
                if descricao_pdf != "-":
                    descricao_pdf = f"{codigo_pix} | {descricao_pdf}"
                else:
                    descricao_pdf = codigo_pix

            # Atualiza APENAS as multas deste grupo (últimas N multas adicionadas)
            # Usa len(indices) para saber quantas multas foram processadas
            quantidade_multas_grupo = len(indices)
            indice_inicio = len(multas_lista) - quantidade_multas_grupo
            
            for j in range(indice_inicio, len(multas_lista)):
                multa = multas_lista[j]
                multa["Órgão Autuador"] = orgao_autuador
                multa["Código de pagamento em barra"] = descricao_pdf
                # Atualiza datas com as do PDF se foram encontradas
                if data_infracao_pdf != "-":
                    multa["Data Infração"] = data_infracao_pdf
                if vencimento_pdf != "-":
                    multa["Data Vencimento"] = vencimento_pdf
        
        return total, multas_lista

    except TimeoutError:
        log("❌ Timeout")
        return 0.0, []
    finally:
        page.close()
        context.close()

# ================= MAIN =================

def main():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        total_geral = 0.0
        todas_multas = []

        for i, v in enumerate(VEICULOS, 1):
            total, multas = processar_veiculo(browser, v, i)
            total_geral += total
            todas_multas.extend(multas)
            if i < len(VEICULOS):
                time.sleep(INTERVALO_ENTRE_CONSULTAS)

        # Salva todas as multas no Excel
        if todas_multas:
            salvar_no_excel(todas_multas)

        log(f"\n💵 TOTAL GERAL: R$ {formatar_valor_br(total_geral)}")
        input("ENTER para sair...")
        browser.close()

if __name__ == "__main__":
    main()
