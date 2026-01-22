import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import os

# ================= LEITURA CSV =================

csv_file = "resultado_detran.csv"

# Verifica se arquivo existe e não está vazio
if not os.path.exists(csv_file) or os.path.getsize(csv_file) == 0:
    print(f"⚠️  Arquivo {csv_file} vazio ou não existe. Criando arquivo com cabeçalho...")
    with open(csv_file, "w", encoding="utf-8") as f:
        f.write("data_hora,placa,renavam,quantidade_multas,valor_total_multas,descricao_multas\n")
    print("✅ Arquivo criado com cabeçalho.")
    exit()

df = pd.read_csv(csv_file, encoding="utf-8-sig", sep=",", engine="python")


df = df.dropna(axis=1, how="all")

# ================= PROCESSAMENTO =================

linhas = []
numero = 0

for _, row in df.iterrows():
    data_hora = row.get("data_hora", "")
    placa = row.get("placa", "")
    renavam = row.get("renavam", "")
    valor_total_csv = row.get("valor_total_multas", 0)

    motivos_raw = str(row.get("motivos_multas", "")).strip()

    if motivos_raw and motivos_raw.lower() != "nenhuma":
        motivos = [m.strip() for m in motivos_raw.split(" | ") if m.strip()]
    else:
        motivos = []

    if not motivos:
        continue

    for motivo in motivos:
        numero += 1
        
        # ================= EXTRAÇÕES =================

        # AIT (letras + números antes do --)
        ait = "-"
        match_ait = re.search(r"([A-Z]{1,3}\d{6,})\s*--", motivo)
        if match_ait:
            ait = match_ait.group(1)

        # AIT Originária
        ait_originaria = "-"

        # Datas
        datas = re.findall(r"\d{2}/\d{2}/\d{4}", motivo)
        
        # Geralmente vem: [vencimento, data_infracao] - vamos inverter
        if len(datas) >= 2:
            try:
                from datetime import datetime
                data1 = datetime.strptime(datas[0], "%d/%m/%Y")
                data2 = datetime.strptime(datas[1], "%d/%m/%Y")
                
                # Se data1 > data2, então data1 é vencimento e data2 é infração
                if data1 > data2:
                    vencimento = datas[0]
                    data_infracao = datas[1]
                else:
                    data_infracao = datas[0]
                    vencimento = datas[1]
            except:
                data_infracao = datas[0]
                vencimento = datas[1]
        elif len(datas) == 1:
            data_infracao = datas[0]
            vencimento = "-"
        else:
            data_infracao = "-"
            vencimento = "-"

        # Valores
        valores = re.findall(r"R\$\s*([\d.,]+)", motivo)
        valor = "-"
        valor_a_pagar = "-"

        if len(valores) == 1:
            valor = f"R$ {valores[0]}"
            valor_a_pagar = f"R$ {valores[0]}"
        elif len(valores) >= 2:
            valor = f"R$ {valores[-2]}"
            valor_a_pagar = f"R$ {valores[-1]}"

        # Motivo/Descrição
        descricao = "-"
        match_desc = re.search(r"--\s*(.+?)\s+\d{2}/\d{2}/\d{4}", motivo)
        if match_desc:
            descricao = match_desc.group(1).strip()
        else:
            match_desc = re.search(r"--\s*(.+)", motivo)
            if match_desc:
                descricao = match_desc.group(1).strip()

        linhas.append({
            "#": numero,
            "AIT": ait,
            "AIT Originária": ait_originaria,
            "Motivo": descricao,
            "Data Infração": data_infracao,
            "Data Vencimento": vencimento,
            "Valor": valor,
            "Valor a Pagar": valor_a_pagar
        })

# ================= DATAFRAME FINAL =================

df_final = pd.DataFrame(linhas)

excel_file = "resultado_detran_organizado.xlsx"
df_final.to_excel(excel_file, index=False, sheet_name="Resultado DETRAN")

# ================= FORMATAÇÃO EXCEL =================

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)

border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = left if cell.column == 5 else center

larguras = {
    "A": 5, "B": 15, "C": 18, "D": 55,
    "E": 14, "F": 14, "G": 16, "H": 16
}

for col, w in larguras.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A2"
wb.save(excel_file)

# ================= RESUMO =================

print(f"\n✅ Excel gerado com sucesso: {excel_file}")
print(f"📊 Total de multas: {len(df_final)}")
